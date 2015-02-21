import tkinter as tk
from tkinter import filedialog
from tkinter.filedialog import askdirectory
import random, win32com, win32api, numpy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.worksheet import Worksheet, header_footer, page, worksheet
from openpyxl.writer.worksheet import write_worksheet

version = '0.9'

class Randomization(tk.Frame):
    def __init__(self,master=None):
        tk.Frame.__init__(self, master)        
        self.grid()
        self.landingScreen()
        
    def landingScreen(self):
        self.button1 = tk.Button(self, text=u'New Rando', command=self.newStudyRando)
        self.button1.grid(column=0, row=0, padx=50, pady=50, sticky='NSEW')
        self.button2 = tk.Button(self, text=u'Outlier Check', command=self.outlierCheckType)
        self.button2.grid(column=2, row=0, padx=35, pady=50, sticky='NSEW')

    def newStudyRando(self):
        
        wn = 372
        hn = 190
        root.geometry('%dx%d+%d+%d' % (wn, hn, x, y))    
        
        def selectDestination():

            self.randoFilepath = tk.filedialog.askdirectory(initialdir='S:/Studies/BRT 2015 Studies', title='Select Study Folder')
            self.randoFilepathLabel.config(text=self.randoFilepath)
        
        self.button1.grid_forget()
        self.button2.grid_forget()
        
        self.randoFilepath = tk.StringVar()
        self.randoFilepath = tk.filedialog.askdirectory(initialdir='S:/Studies/BRT 2015 Studies', title='Select Study Folder')

        self.studyNumber = tk.StringVar()
        self.entry1 = tk.Entry(self, textvariable=self.studyNumber)
        self.entry1.grid(column=0, row=0, padx=8, pady=8, sticky='EW')
        self.studyNumber.set(u'Enter Study Number')
        self.entry1.focus()
        self.entry1.select_range(0, tk.END)
        
        self.startAnimalNum = tk.IntVar()
        self.entry2 = tk.Entry(self, textvariable=self.startAnimalNum)
        self.entry2.grid(column=0, row=1, padx=8, pady=8, sticky='EW')
        self.startAnimalNum.set(u'Enter Starting Animal Number')
        
        self.totalAnimals = tk.IntVar()
        self.entry3 = tk.Entry(self, text=self.totalAnimals)
        self.entry3.grid(column=0, row=2, padx=8, pady=8, sticky='EW')
        self.totalAnimals.set(u'Enter Total Number of Animals')
        
        self.v = tk.IntVar()

        self.radio1 = tk.Radiobutton(self, text="Rats", variable=self.v, value=1)
        self.radio1.grid(column=0, row=3, padx=8, pady=8, sticky='W')
        self.radio2 = tk.Radiobutton(self, text="Mice", variable=self.v, value=2)
        self.radio2.grid(column=0, row=3, padx=8, pady=8)
        
        self.randoFilepathLabel = tk.Label(self, text=self.randoFilepath)
        self.randoFilepathLabel.grid(column=0, row=4, padx=8, pady=8, sticky='EW')
        
        button3 = tk.Button(self, text=u'Randomize!', command=self.onRandoClick)
        button3.grid(column=1, row=2, padx= 10, sticky='NSEW')
        
        button6 = tk.Button(self, text=u'Select Destination', command=selectDestination)
        button6.grid(column=1, row=4, padx= 10, sticky='NSEW')
        
        self.grid_columnconfigure((0), weight=1, minsize=250)
        self.grid_columnconfigure((1), weight=1, minsize=125)
        
    def onRandoClick(self):

        ##Init
        
        anType = ''
        if self.v == 1:
            anType = 'Rat'
        else:
            anType = 'Mouse'
        count = 0
        data = []
        header = ["%s #" % anType, "Weight (g)"]
        endRow = self.totalAnimals.get() + 1
        wtNumFormat = 'General'
        
        if self.v == 1:
            wtNumFormat = '0'
        else:
            wtNumFormat = '0.0'
        
        headerStyle = Style(font=Font(bold=True),
                     border=Border(left=Side(border_style='thin',
                                color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     outline=Side(border_style='thin',
                                  color='FF000000'),
                     vertical=Side(border_style='hair',
                                   color='FF000000'),
                     horizontal=Side(border_style='hair',
                                     color='FF000000')),
                     alignment=Alignment(horizontal='center',
                                         vertical='center',
                                        wrap_text=True))
        
        numColumnStyle = Style(border=Border(left=Side(border_style='thin',
                                color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     outline=Side(border_style='thin',
                                  color='FF000000'),
                     vertical=Side(border_style='hair',
                                   color='FF000000'),
                     horizontal=Side(border_style='hair',
                                     color='FF000000')),
                     alignment=Alignment(horizontal='center',
                            vertical='center'),
                     number_format='General',
                     protection=Protection(locked=True,
                                   hidden='inherit'))
        
        wtColumnStyle = Style(border=Border(left=Side(border_style='thin',
                                color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     outline=Side(border_style='thin',
                                  color='FF000000'),
                     vertical=Side(border_style='hair',
                                   color='FF000000'),
                     horizontal=Side(border_style='hair',
                                     color='FF000000')),
                     alignment=Alignment(horizontal='center',
                            vertical='center'),
                     number_format=wtNumFormat,
                     protection=Protection(locked=True,
                                   hidden='inherit'))
        

        while count < self.totalAnimals.get():
            data.append(self.startAnimalNum.get() + count)
            count += 1

        random.shuffle(data)
        
        ##Initialize workbook creation
        
        wb = Workbook()
        ws = wb.active
        ws.title = '%s' % str(self.studyNumber.get())
        wshf = ws.header_footer
        wspm = ws.page_margins
        
        ##Column Headers
        
        ws.cell('A1').value = header[0]
        ws.cell('B1').value = header[1]
        ws.cell('A1').style = headerStyle
        ws.cell('B1').style = headerStyle
        
        ##Rando sign-off box
        
        ws.merge_cells('E3:F7')
        
        for row in ws.iter_rows('E3:F7'):
            for cell in row: 
                cell.style = headerStyle
                
        ws.cell('E3').value = 'Randomized using BRT Randomization v%s (Date/Initial) _________________' % version
        ws.cell('E3').style = headerStyle
        
        ##Balance Info Box
        
        ws.merge_cells('D10:G10')
        ws.merge_cells('D11:E11')
        ws.merge_cells('D12:E12')
        ws.merge_cells('D13:E13')
        ws.merge_cells('D14:E14')
        ws.merge_cells('D15:E15')
        ws.merge_cells('D16:E16')
        ws.merge_cells('F11:G11')
        ws.merge_cells('F12:G12')
        ws.merge_cells('F13:G13')
        ws.merge_cells('F14:G14')
        ws.merge_cells('F15:G15')
        ws.merge_cells('F16:G16')
        
        for row in ws.iter_rows('D10:G16'):
            for cell in row: 
                cell.style = headerStyle
        
        ws.cell('D10').value = 'Balance Info'
        ws.cell('D11').value = 'Balance Used'
        ws.cell('D12').value = 'Serial Number'
        ws.cell('D13').value = 'Calibrated'
        ws.cell('D14').value = 'Verified Before'
        ws.cell('D15').value = 'Verified After'
        ws.cell('D16').value = 'Calibration Due'


        ##Write shuffled values to cells
        
        j = 0
        
        for i in range(2, self.totalAnimals.get() + 2):
            ws.cell(row = i, column = 1).value = data[j]
            j = j + 1
        
        wb.create_named_range('data', ws, 'A2:B%s' % str(j + 1))
        
        ##Set styling for randomized data table
        
        for row in ws.iter_rows('A2:A%r' % endRow):
            for cell in row: 
                cell.style = numColumnStyle
        
        for row in ws.iter_rows('B2:B%r' % endRow):
            for cell in row: 
                cell.style = wtColumnStyle
                
        ##Add headers and footers
        
        wshf.left_header.text = 'Date: _______________'
        wshf.center_header.text = str(self.studyNumber.get()) + '\rRandomization'
        wshf.right_header.text = 'Operator: ________________\r________________\r________________\r________________'
        
#        wshf.left_footer.text = 'Randomized using BRT Randomization\rversion 1.0 \r(Date/Initial) _________________'
        wshf.center_footer.text = 'Page &P of &N\r\r'

        
        ##Set print titles, margins, and column widths
        
        ws.add_print_title(1, rows_or_cols='rows')
        
        wspm.left = .75
        wspm.right = .75
        wspm.top = 1.2
        wspm.bottom = 1.4
        wspm.header = .5
        wspm.footer = .5
        
        ws.cell('C1').value = ''
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 11
        ws.column_dimensions['E'].width = 11
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 11
        
        ##Enable sheet protection and save the file
        
#        ws.protection.enable()
        wb.save(self.randoFilepath + '//%s Rando.xlsx' % self.studyNumber.get())
        
        win32api.MessageBox(0, 'Success!', '%s Randomization' % self.studyNumber.get())
        
        root.destroy()
        
        
    def outlierCheckType(self):
       
        self.button1.grid_forget()
        self.button2.grid_forget()
        
        self.button4 = tk.Button(self, text=u'Mean \u00B1 20%', command=lambda: self.outlierCheck('Percent'))
        self.button4.grid(column=0, row=0, padx=50, pady=50, sticky='NSEW')
        self.button5 = tk.Button(self, text=u'Mean \u00B1 3 x StDev', command=lambda: self.outlierCheck('Range'))
        self.button5.grid(column=1, row=0, padx=10, pady=50, sticky='NSEW')

    def outlierCheck(self, ocType):
    
        self.button4.grid_forget()
        self.button5.grid_forget()
        
        defaultStyle = Style(font=Font(name='Calibri',
                 size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000'),
                fill=PatternFill(fill_type=None,
                 start_color='FFFFFFFF',
                 end_color='FF000000'),
                border=Border(left=Side(border_style=None,
                                color='FF000000'),
                      right=Side(border_style=None,
                                 color='FF000000'),
                      top=Side(border_style=None,
                               color='FF000000'),
                      bottom=Side(border_style=None,
                                  color='FF000000'),
                      diagonal=Side(border_style=None,
                                    color='FF000000'),
                      diagonal_direction=0,
                      outline=Side(border_style=None,
                                   color='FF000000'),
                      vertical=Side(border_style=None,
                                    color='FF000000'),
                      horizontal=Side(border_style=None,
                                     color='FF000000')),
                alignment=Alignment(horizontal='general',
                         vertical='bottom',
                         text_rotation=0,
                         wrap_text=False,
                         shrink_to_fit=False,
                         indent=0),
                number_format='General',
                protection=Protection(locked='inherit',
                           hidden='inherit'))
        
        headerStyle = Style(font=Font(bold=True),
                     border=Border(left=Side(border_style='thin',
                                color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     outline=Side(border_style='thin',
                                  color='FF000000'),
                     vertical=Side(border_style='hair',
                                   color='FF000000'),
                     horizontal=Side(border_style='hair',
                                     color='FF000000')),
                     alignment=Alignment(horizontal='center',
                                         vertical='center',
                                        wrap_text=True))
        
        file_path = ''
        
        file_path = tk.filedialog.askopenfilename(initialdir='S:\\Studies\\', title='Select Rando File', parent=root)

        animalNums = []
        animalWts = []
        onlyWts = []
        outliers = {}
        outlierAnNums = []
        lowrange = int
        highrange = int
        roundPlaces = int
        ocTypeString = str
        anType = ''
        
        wb = load_workbook(filename = file_path)
        file_path = file_path[:-5] + ' Outlier Check.xlsx'
        wb.save(file_path)
        ws = wb.active
        wshf = ws.header_footer
        
        anType = ws.cell('A1').value.split(' ')
        anType = anType[1]
        
        if anType == 'Rat':
            roundPlaces = 0
        else:
            roundPlaces = 1
        
        data = wb.get_named_range('data').value
        split = data.split('$')               ##Create named range of data and get max row from that
        maxDataRow = split[4]
        
        for row in ws.iter_rows('A2:A%s' % maxDataRow):
            for an in row:                                  ##Need to skip rows without weights
                animalNums.append(an.value)
                
        for row in ws.iter_rows('B2:B%s' % maxDataRow):
            for wt in row:
                animalWts.append(wt.value)
                    
        onlyWts = [x for x in animalWts if x != None]
        
        dataDict = dict(zip(animalNums, animalWts))
        wtsFirstDict = dict(zip(animalWts, animalNums))
        
        print(ws.cell('B2').value)
        
        totAnimals = int(maxDataRow) - 1
        wtMean = round(numpy.mean(onlyWts), roundPlaces)
        wtStdev = round(numpy.std(onlyWts), roundPlaces)
        plus20pct = round(wtMean * 1.2, roundPlaces)
        minus20pct = round(wtMean * 0.8, roundPlaces)
        plus3std = round(wtMean + 3 * wtStdev, roundPlaces)
        minus3std = round(wtMean - 3 * wtStdev, roundPlaces)

        
        if ocType == 'Percent':
            highrange = plus20pct
            lowrange = minus20pct
            ocTypeString = u'Mean \u00B1 20%'
            for anNum, anWt in dataDict.items():
                if anWt == None:
                    continue
                elif (anWt < minus20pct) or (anWt > plus20pct):
                    outliers[anNum] = anWt
        
        if ocType == 'Range':
            highrange = plus3std
            lowrange = minus3std
            ocTypeString = u'Mean \u00B1 3xStDev'
            for anNum, anWt in dataDict.items():
                if anWt == None:
                    continue
                elif (anWt < minus3std) or (anWt > plus3std):
                    outliers[anNum] = anWt
        
        if len(outliers) == 0:
            outlierAnNums = []
        else:
            for b in outliers:
                outlierAnNums.append(str(b))
                
        if len(outlierAnNums) == 0:
            outlierStatement = 'There are no weight outliers'
        elif len(outlierAnNums) == 1:
            outlierStatement = 'Animal number %s is a weight outlier' % ', '.join(outlierAnNums)
        else:
            outlierStatement = 'Animal numbers %s are weight outliers' % ', '.join(outlierAnNums)    
            
        
        ws.unmerge_cells('E3:F7')
        ws.unmerge_cells('D15:E15')
        ws.unmerge_cells('D16:E16')
        ws.unmerge_cells('F15:G15')
        ws.unmerge_cells('F16:G16')
        
        wshf.left_header.text = ''
        wshf.center_header.text = str(ws.title) + '\rOutlier Check'
        wshf.right_header.text = ''
        
        for row in ws.iter_rows('E3:F7'):
            for cell in row: 
                cell.style = defaultStyle
                
        ws.cell('E3').value = ''
        ws.cell('E3').style = defaultStyle
        
        ws.cell('D10').value = 'Outlier Check   (%s)' %ocTypeString
        ws.cell('D11').value = 'n: '
        ws.cell('D12').value = 'Mean:'
        ws.cell('D13').value = 'Std Dev:'
        ws.cell('D14').value = 'Range:'
        ws.cell('D15').value = outlierStatement
        ws.cell('F11').value = totAnimals
        ws.cell('F12').value = '%s g' % str(wtMean)
        ws.cell('F13').value = '%s g' % str(wtStdev)
        ws.cell('F14').value = '%s g --> %s g' % (str(lowrange), str(highrange))
                
        ws.merge_cells('D15:G16')
        
        for row in ws.iter_rows('D10:G16'):
            for cell in row: 
                cell.style = headerStyle
                
        wb.save(file_path)
        
        win32api.MessageBox(0, 'Success!', '%s Outlier Check' % ws.title)
        
        root.destroy()
            
if __name__ == '__main__':
    root = tk.Tk()
    root.title('BRT Randomization v%s' % version)
    
    #make my screen dimensions work
    w = 340 #The value of the width
    h = 130 #The value of the height of the window

    # get screen width and height
    ws = root.winfo_screenwidth()#This value is the width of the screen
    hs = root.winfo_screenheight()#This is the height of the screen

    # calculate position x, y
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)

    #This is responsible for setting the dimensions of the screen and where it is
    #placed
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    
    app = Randomization(master=root)
    app.mainloop()