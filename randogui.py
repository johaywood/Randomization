import tkinter as tk
from tkinter import filedialog
import xl, csv, random
import win32com
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.worksheet import Worksheet, header_footer, page, worksheet
from openpyxl.writer.worksheet import write_worksheet

class Randomization(tk.Frame):
    def __init__(self,master=None):
        tk.Frame.__init__(self, master)        
        self.grid()
        self.landingScreen()
        
    def landingScreen(self):
        self.button1 = tk.Button(self, text=u'New Rando', command=self.newStudyRando)
        self.button1.grid(column=0, row=1, padx=50, pady=50, sticky='NSEW')
        self.button2 = tk.Button(self, text=u'Outlier Check', command=self.outlierCheckType)
        self.button2.grid(column=2, row=1, padx=50, pady=50, sticky='NSEW')

    def newStudyRando(self):
        
        def selectDestination():

            self.randoFilepath = tk.filedialog.askdirectory()
            self.randoFilepathLabel.config(text=self.randoFilepath)
        
        self.button1.grid_forget()
        self.button2.grid_forget()
        
        self.randoFilepath = tk.StringVar()
        self.randoFilepath = tk.filedialog.askdirectory()

        self.studyNumber = tk.StringVar()
        self.entry1 = tk.Entry(self, textvariable=self.studyNumber)
        self.entry1.grid(column=0, row=0, padx=8, pady=8, sticky='EW')
        self.studyNumber.set(u'Enter Study Number')
        
        self.startAnimalNum = tk.IntVar()
        self.entry2 = tk.Entry(self, textvariable=self.startAnimalNum)
        self.entry2.grid(column=0, row=1, padx=8, pady=8, sticky='EW')
        self.startAnimalNum.set(u'Enter Starting Animal Number')
        
        self.totalAnimals = tk.IntVar()
        self.entry3 = tk.Entry(self, text=self.totalAnimals)
        self.entry3.grid(column=0, row=2, padx=8, pady=8, sticky='EW')
        self.totalAnimals.set(u'Enter Starting Animal Number')
        
        self.randoFilepathLabel = tk.Label(self, text=self.randoFilepath)
        self.randoFilepathLabel.grid(column=0, row=4, padx=8, pady=8, sticky='EW')
        
        button3 = tk.Button(self, text=u'Randomize!', command=self.onRandoClick)
        button3.grid(column=1, row=1, padx= 10, sticky='NSEW')
        
        button6 = tk.Button(self, text=u'Select Destination', command=selectDestination)
        button6.grid(column=1, row=4, padx= 10, sticky='NSEW')
        
        self.grid_columnconfigure((0), weight=1, minsize=250)
        self.grid_columnconfigure((1), weight=1, minsize=125)
        
    def onRandoClick(self):

        count = 0
        data = []
        header = ["Animal #", "Weight (g)"]
        endRow = self.totalAnimals.get() + 1
        headerStyle = Style(font=Font(bold=True),
                     border=Border(left=Side(border_style='thin',
                                color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     outline=Side(border_style='thin',
                                  color='FF000000'),
                     vertical=Side(border_style='hair',
                                   color='FF000000'),
                     horizontal=Side(border_style='hair',
                                     color='FF000000')),
                     alignment=Alignment(horizontal='center',
                                         vertical='center'))
        columnStyle = Style(border=Border(left=Side(border_style='thin',
                                color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     outline=Side(border_style='thin',
                                  color='FF000000'),
                     vertical=Side(border_style='hair',
                                   color='FF000000'),
                     horizontal=Side(border_style='hair',
                                     color='FF000000')),
                     alignment=Alignment(horizontal='center',
                            vertical='center'),
                     number_format='General',
                     protection=Protection(locked=True,
                                   hidden='inherit'))

        while count < self.totalAnimals.get():
            data.append(self.startAnimalNum.get() + count)
            count += 1

        random.shuffle(data)
        
        wb = Workbook()
        ws = wb.active
        wshf = ws.header_footer
        
        ws.cell('A1').value = header[0]
        ws.cell('B1').value = header[1]
        ws.cell('A1').style = headerStyle
        ws.cell('B1').style = headerStyle
        
        j = 0
        
        for i in range(2, self.totalAnimals.get() + 2):
            ws.cell(row = i, column = 1).value = data[j]
            j = j + 1
        
        for row in ws.iter_rows('A2:B%r' % endRow):
            for cell in row: 
                cell.style = columnStyle
                
        wshf.left_header.text = 'Date: _______________'
        wshf.center_header.text = str(self.studyNumber.get()) + '\rRandomization'
        wshf.right_header.text = 'Operator: ________________\r________________\r________________\r________________'
        
        wshf.left_footer.text = 'Randomized using BRT Randomization\rversion 1.0 \r(Date/Initial) _________________'
        wshf.center_footer.text = 'Page &P of &N'
        wshf.right_footer.text = '&[Path]'
        
        wb.create_named_range('_xlnm.Print_Titles', ws, '$1:$1', self)
        
        print(wb._named_ranges)
        
        #ws.PageMargins(left=0.75, right=0.75, top=1.114, bottom=1.416, header=0.5, footer=0.5)
        
        ws.protection.enable()
        wb.save(self.randoFilepath + '//%s Randomization.zip' % self.studyNumber.get())
#        
#        xl = win32com.client.Dispatch('Excel.application')
#        wb = xl.Workbooks.Add()
#        sheet = wb.Sheets(1)
#        sheet.Name='toto'
#        xl.ActiveWorkbook.Names.Add(Name='_xlnm.Print_Titles', RefersTo=str(sheet.Name + '$1:$1'))
        
        #root.destroy()
        
    def outlierCheckType(self):
       
        self.button1.grid_forget()
        self.button2.grid_forget()
        
        self.button4 = tk.Button(self, text=u'20% of Mean', command=lambda: self.outlierCheck('Percent'))
        self.button4.grid(column=0, row=0, padx=50, pady=50, sticky='NSEW')
        self.button5 = tk.Button(self, text=u'Mean +- 3 StDev', command=lambda: self.outlierCheck('Range'))
        self.button5.grid(column=1, row=0, padx=50, pady=50, sticky='NSEW')

    def outlierCheck(self, ocType):
    
        self.button4.grid_forget()
        self.button5.grid_forget()
        
        self.file_path = ''
        
        self.file_path = tk.filedialog.askopenfilename()
        
        self.file_contents = []
        
        with open(self.file_path, 'r') as of:
            b = csv.reader(of, delimiter=',')
            for row in b:
                self.file_contents.append(row)
        
        lastRow = len(self.file_contents)
        meanCell = len(self.file_contents) + 2
        stDevCell = len(self.file_contents) + 3
        outlierCheckRows = []
        
        outlierCheckRows.append(['','',''])
        outlierCheckRows.append(['Mean', '=average($B$2:$B$%r)' % lastRow])
        outlierCheckRows.append(['StDev', '=stdev($B$2:$B$%r)' % lastRow])
        if ocType == 'Percent':
            outlierCheckRows.append(['Range (20% of Mean)', '', '=$B$%r*0.8' % meanCell, 'to', '=$B$%r*1.2' % meanCell])
        elif ocType == 'Range':
            outlierCheckRows.append(['Range (+- 3xStDev)', '', '=$B$%r-3*$B$%r' % (meanCell, stDevCell), 'to', '=$B$%r-3*$B$%r' % (meanCell, stDevCell)])
              
        with open(self.file_path, 'a', newline='') as fp:   
            a = csv.writer(fp, delimiter=',')
            a.writerows(outlierCheckRows)
            
        root.destroy()
            
if __name__ == '__main__':
    root = tk.Tk()
    app = Randomization(master=root)
    app.mainloop()